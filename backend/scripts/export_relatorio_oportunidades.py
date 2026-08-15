"""Scraper Playwright — relatório "jmarques_oportunidades_notas" do eGO CRM
(módulo Oportunidades). Mesmo mecanismo que `export_relatorio_imoveis.py`
(login httpx + cookies injectadas, popup Relatórios, download) — ver esse
ficheiro para o detalhe do porquê de cada passo.

Escreve em `teste_oportunidades` — mesmos nomes de coluna que `oportunidades`
(tabela real de produção, ~90 colunas, confirmada 2026-07-27 mas nunca
documentada), todas `text` (staging) + `extra` jsonb p/ cabeçalhos do
relatório sem correspondência. Relatório devolve 1 linha por (oportunidade
× nota) — mesmo padrão do relatório de Imóveis — `_merge_notas` agrupa por
`oportunidade_ref` (cabeçalho "Referência", a 1ª ocorrência — a 2ª,
"Referência (2)", é o `imovel_ref` associado).

Filtros aplicados (pedido do utilizador, 2026-07-27): desseleccionar
"Minhas oportunidades" (filtro por defeito, seleccionado ao carregar a
página) + "Editado em > Últimas 48 horas" — não existia pesquisa
personalizada "editado à 3 dias"
gravada no eGO, confirmado ao vivo via Playwright.

Correr a partir de backend/: python scripts/export_relatorio_oportunidades.py
Depende de backend/scripts/requirements-scraper.txt (playwright, openpyxl)
+ `python -m playwright install chromium` (uma vez).
"""
import asyncio
import csv
import datetime
import io
import re
import sys
import unicodedata
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from playwright.async_api import TimeoutError as PlaywrightTimeout
from playwright.async_api import async_playwright

from app.config import settings
from app.db.supabase_client import get_supabase
from app.integrations import egorealestate_crm

REPORT_NAME = "jmarques_oportunidades_notas"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# Mesmos nomes que `oportunidades` (migration 0011) — sem `id`/`extra`/`criado_em`.
_KNOWN_COLUMNS = {
    "oportunidade_ref", "tipo_oportunidade", "data_criacao_raw", "data_criacao_iso",
    "cliente_nome", "cliente_telefone", "cliente_email", "url", "origem_lista",
    "titulo_imovel", "backup_em", "atualizado_em", "etapa_atual", "etapa_dias",
    "preferencia_imovel", "oportunidade_estado", "ultima_consulta_em", "responsavel",
    "xlsx_ref_im", "xlsx_angariador", "xlsx_valor", "xlsx_comm_ag", "xlsx_comm_ang",
    "xlsx_comm_vend", "xlsx_margem", "xlsx_origem", "xlsx_estado_prop", "xlsx_data_fecho",
    "xlsx_data_prop", "xlsx_exclusivo", "xlsx_atualizado_em", "ego_data_criacao",
    "ego_editado_em", "tipo_pedido", "origem", "sub_origem", "portal", "imovel_preco",
    "proposta", "diferenca", "data_proposta", "estado_proposta", "etapa_proposta",
    "probabilidade", "agencia", "agencia_email", "agencia_telefone", "etapas_cpcv",
    "etapas_escrituras", "checklist_financiamento", "checklist_cpcv", "checklist_escrituras",
    "imovel_proprietario", "imovel_ref", "imovel_natureza", "imovel_distrito",
    "imovel_concelho", "imovel_freguesia", "imovel_venda", "imovel_arrendamento",
    "estado_fechado_em", "data_escritura", "valor_negocio", "equipa_responsavel",
    "visita_imovel_ref", "visita_anulada", "visita_interessado", "visita_data",
    "visita_imovel_proprietario", "visita_cliente", "visita_pontos_positivos",
    "visita_pontos_negativos", "visita_sobre_negocio", "visita_observacoes",
    "visita_responsavel", "imovel2_ref", "imovel2_distrito", "imovel2_concelho",
    "imovel2_freguesia", "imovel2_venda", "imovel2_arrendamento", "pref_tipologia",
    "pref_orcamento_max", "pref_zona", "pref_extraido_em", "pref_outros", "ponto_situacao",
    "ponto_situacao_alterado_em", "motivo_transacao", "pref_natureza", "pref_negocio",
    "pref_preco_min", "pref_disponibilidade", "visita_ref_ego",
}

# Cabeçalhos reais do relatório "jmarques_oportunidades_notas" (visto
# 2026-07-27) que não batem por normalização simples.
_ALIASES = {
    "referencia": "oportunidade_ref",
    "referencia_2": "imovel_ref",
    "estado": "oportunidade_estado",
    "etapa": "etapa_atual",
    "venda": "imovel_venda",
    "editado_em": "ego_editado_em",
    "data_de_criacao": "ego_data_criacao",
    "proprietario": "imovel_proprietario",
    "tipo_de_pedido": "tipo_pedido",
    "potencial_cliente": "cliente_nome",
    "tipo_de_negocio": "tipo_oportunidade",
    "link": "url",
    "concelho": "imovel_concelho",
    "distrito": "imovel_distrito",
    "freguesia": "imovel_freguesia",
    "natureza": "imovel_natureza",
    "ponto_de_situacao": "ponto_situacao",
    "ponto_de_situacao_alterado_em": "ponto_situacao_alterado_em",
    "motivo_de_transacao": "motivo_transacao",
    "checklist_de_cpcv": "checklist_cpcv",
    "checklist_de_escrituras": "checklist_escrituras",
    "checklist_de_financiamento": "checklist_financiamento",
    "etapas_de_cpcv": "etapas_cpcv",
    "etapas_de_escrituras": "etapas_escrituras",
    "data_da_escritura": "data_escritura",
    "data_da_proposta": "data_proposta",
    "valor_do_negocio": "valor_negocio",
    "equipa_do_responsavel": "equipa_responsavel",
}

# O relatório devolve 1 linha por (oportunidade × nota) — dados da
# oportunidade repetem-se idênticos em cada linha, só estes campos mudam
# por nota. "Descrição (2)" é o texto da nota em si (distinto de
# "Descrição", a descrição da própria oportunidade).
_NOTA_FIELDS = ("Criado em", "Criado por", "Tipo de nota", "Descrição (2)", "Anexos", "Apagado em", "Apagado por")


async def _httpx_cookies() -> list[dict]:
    async with egorealestate_crm.authenticated_client() as client:
        await egorealestate_crm._login(client)
        return [
            {"name": c.name, "value": c.value, "domain": c.domain, "path": c.path or "/"}
            for c in client.cookies.jar
        ]


def _json_safe(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def _load_workbook_safe(path: Path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except IndexError:
        # eGO exporta xlsx com `styles.xml` a referenciar cellStyleXfs fora de
        # gama (quirk comum de geradores .NET) — ver export_relatorio_imoveis.py.
        patched = path.with_suffix(".patched.xlsx")
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*>.*?</(\w+:)?cellStyles>", b"", data, flags=re.DOTALL)
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*/>", b"", data)
                zout.writestr(item, data)
        return openpyxl.load_workbook(patched, data_only=True)


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
        result.append(h if seen[h] == 1 else f"{h} ({seen[h]})")
    return result


def _normalize(header: str) -> str:
    text = unicodedata.normalize("NFKD", header).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()


def _data_iso(valor: str) -> str | None:
    """"Data de criação" vem "dd/mm/yyyy" — `data_criacao_iso` (produção,
    populada por outro processo) usa "yyyy-mm-dd"; converter em vez de
    deixar vazio só porque não temos o texto scrapeado original completo
    que alimenta `data_criacao_raw`."""
    try:
        return datetime.datetime.strptime(valor.strip(), "%d/%m/%Y").date().isoformat()
    except (ValueError, AttributeError):
        return None


def _map_row(row: dict) -> dict:
    record: dict = {"extra": {}}
    for header, value in row.items():
        key = _normalize(str(header))
        key = _ALIASES.get(key, key)
        if key in _KNOWN_COLUMNS and key not in record:
            record[key] = None if value is None else str(value)
        else:
            record["extra"][str(header)] = value

    if record.get("ego_data_criacao"):
        iso = _data_iso(record["ego_data_criacao"])
        if iso:
            record["data_criacao_iso"] = iso
    return record


def _merge_notas(records: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    order: list[str] = []
    ignorados = 0
    for record in records:
        ref = record.get("oportunidade_ref")
        if not ref:
            ignorados += 1
            continue
        core = {k: v for k, v in record.items() if k != "extra"}
        if ref not in grouped:
            grouped[ref] = dict(core)
            grouped[ref]["extra"] = {k: v for k, v in record["extra"].items() if k not in _NOTA_FIELDS}
            grouped[ref]["extra"]["notas"] = []
            order.append(ref)
        else:
            # só avisar se os dados fora dos campos de nota realmente
            # diferirem — repetição normal (mais uma nota) não é problema.
            existente = {k: v for k, v in grouped[ref].items() if k != "extra"}
            if core != existente:
                print(f'  aviso: "{ref}" repetido no relatório com dados diferentes fora dos campos de nota — mantida só a 1ª ocorrência')
        nota = {k: record["extra"].get(k) for k in _NOTA_FIELDS}
        if any(v is not None for v in nota.values()):
            grouped[ref]["extra"]["notas"].append(nota)
    if ignorados:
        print(f"  {ignorados} linha(s) sem oportunidade_ref ignoradas")
    return [grouped[ref] for ref in order]


def _parse_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".xlsx":
        wb = _load_workbook_safe(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = _dedupe_headers([str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])])
        return [{k: _json_safe(v) for k, v in zip(header, row)} for row in rows[1:]]
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    raise ValueError(f"Formato não suportado: {path.suffix}")


async def _click_sidetag(page, texto: str) -> bool:
    return await page.evaluate(
        """(texto) => {
            const span = Array.from(document.querySelectorAll('span.sideTag'))
                .find(e => e.textContent.trim() === texto);
            if (!span) return false;
            const link = span.querySelector('a') || span;
            link.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
            return true;
        }""",
        texto,
    )


_MINHAS = "Minhas oportunidades"

# A tag aparece DUAS vezes quando está aplicada (a opção e o chip de filtro
# activo) e é na segunda que se clica.
_JS_CONTA = (
    "() => Array.from(document.querySelectorAll('span.sideTag a'))"
    f".filter(e => e.textContent.trim() === '{_MINHAS}').length"
)


async def _desseleccionar_minhas(page) -> None:
    """Igual ao de `scraper/oportunidades_completo.py` — ver lá o porquê.

    Duplicado de propósito: o `scraper/` é uma app Fly.io à parte, com o seu
    próprio Dockerfile e requirements, e não importa nada de `backend/`.
    Partilhar isto exigiria um pacote comum por 40 linhas.
    """
    try:
        await page.wait_for_function(f"({_JS_CONTA})() >= 2", timeout=15000)
    except PlaywrightTimeout:
        diag = await page.evaluate(
            """() => {
                const ancoras = Array.from(document.querySelectorAll('span.sideTag a'));
                return {
                    sideTags: document.querySelectorAll('span.sideTag').length,
                    rotulos: ancoras.slice(0, 10).map(e => e.textContent.trim()),
                };
            }"""
        )
        encontrados = await page.evaluate(f"({_JS_CONTA})()")
        raise RuntimeError(
            f'"{_MINHAS}" não encontrado para desseleccionar após 15s '
            f"(encontrados {encontrados}, precisos 2). "
            f"span.sideTag na página: {diag['sideTags']}. "
            f"Primeiros rótulos: {diag['rotulos']}"
        )

    await page.evaluate(
        """() => {
            const els = Array.from(document.querySelectorAll('span.sideTag a'))
                .filter(e => e.textContent.trim() === 'Minhas oportunidades');
            els[1].dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
        }"""
    )


async def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    print("A obter sessão autenticada (httpx)...")
    cookies = await _httpx_cookies()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False, slow_mo=400)
        context = await browser.new_context(accept_downloads=True)
        await context.add_cookies(cookies)
        page = await context.new_page()

        download_path: Path | None = None
        download_done = asyncio.Event()

        async def _on_download(download):
            nonlocal download_path
            download_path = OUTPUT_DIR / download.suggested_filename
            await download.save_as(download_path)
            print(f"Download: {download_path}")
            download_done.set()

        context.on("download", _on_download)

        print("A navegar para Oportunidades...")
        await page.goto(
            f"{settings.egorealestate_crm_base_url}/egocore/leads",
            wait_until="networkidle",
            timeout=30000,
        )
        if "login" in page.url:
            raise RuntimeError("Sessão não autenticou — cookies inválidas ou login mudou.")
        await page.wait_for_timeout(2000)

        # "Minhas oportunidades" vem seleccionado por defeito (filtro por
        # utilizador logado). Clicar em "Todas as oportunidades" não resolve
        # (testado ao vivo, dá 0 resultados) — tem de se DESSELECCIONAR
        # "Minhas oportunidades" clicando na própria tag já seleccionada,
        # mesmo padrão do script antigo (JMWAIweb export-oportunidades.js).
        print('A desseleccionar "Minhas oportunidades"...')
        await _desseleccionar_minhas(page)
        await page.wait_for_timeout(2000)

        print('A aplicar filtro "Editado em > Últimas 48 horas"...')
        if not await _click_sidetag(page, "Últimas 48 horas"):
            raise RuntimeError('Filtro "Últimas 48 horas" não encontrado.')
        await page.wait_for_timeout(2000)

        print("A limpar selecção...")
        await page.evaluate("document.querySelector('.btDeselectAll, a.btDeselectAll')?.click()")
        await page.wait_for_timeout(1000)

        print("A seleccionar todos...")
        await page.evaluate("document.querySelector('.btSelectAll, a.btSelectAll')?.click()")
        await page.wait_for_timeout(1500)

        await page.evaluate(
            """() => {
                const ok = Array.from(document.querySelectorAll('button, a'))
                    .find(b => b.textContent.trim() === 'OK' || b.textContent.trim() === 'Fechar');
                if (ok) ok.click();
            }"""
        )
        await page.wait_for_timeout(500)

        print("A abrir Relatórios...")
        clicked_rel = await page.evaluate(
            """() => {
                const el = Array.from(document.querySelectorAll('a'))
                    .find(e => e.textContent.trim() === 'Relatórios' && e.getAttribute('onclick')?.includes('popupreports'));
                if (el) { el.click(); return true; }
                return false;
            }"""
        )
        if not clicked_rel:
            raise RuntimeError('Link "Relatórios" não encontrado — selecção pode ter falhado (0 resultados?).')
        await page.wait_for_timeout(3000)

        print(f'A clicar no relatório "{REPORT_NAME}"...')
        clicked_report = await page.evaluate(
            """(nome) => {
                const items = Array.from(document.querySelectorAll('#ReportsList .popupReportItem, #ReportsPopup .popupReportItem'));
                const item = items.find(el => el.textContent.trim() === nome);
                if (item) { (item.querySelector('a') || item).click(); return true; }
                return false;
            }""",
            REPORT_NAME,
        )
        if not clicked_report:
            raise RuntimeError(f'Relatório "{REPORT_NAME}" não encontrado na lista de relatórios gravados.')

        try:
            await asyncio.wait_for(download_done.wait(), timeout=30)
        except asyncio.TimeoutError:
            pass

        await browser.close()

    if not download_done.is_set() or download_path is None:
        raise RuntimeError("Nenhum download detectado em 30s.")

    print(f"A parsear {download_path.name}...")
    rows = _parse_rows(download_path)
    print(f"{len(rows)} linhas no relatório.")
    if not rows:
        return

    registos = _merge_notas([_map_row(row) for row in rows])
    print(f"{len(registos)} oportunidades únicas (notas fundidas).")
    print("A gravar em teste_oportunidades...")
    supabase = get_supabase()
    for i in range(0, len(registos), 100):
        supabase.table("teste_oportunidades").insert(registos[i : i + 100]).execute()
    print(f"{len(registos)} linhas gravadas em teste_oportunidades.")
    sem_match = {k for r in registos for k in r["extra"]}
    if sem_match:
        print(f"Colunas do relatório sem correspondência em `oportunidades` (foram para `extra`): {sorted(sem_match)}")


if __name__ == "__main__":
    asyncio.run(main())
