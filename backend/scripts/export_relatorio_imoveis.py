"""Scraper Playwright — relatório "jmarques_imoveis_notas" do eGO CRM (módulo Imóveis).

Só local — Chromium headless precisa de mais RAM do que a VM Fly.io actual
(256MB) tem, correr lá dentro arriscava OOM no mesmo backend que trata
voz/WhatsApp. Reaproveita o login httpx já testado (`egorealestate_crm.py`),
injecta as cookies numa sessão Playwright em vez de reimplementar o
preenchimento do formulário de login.

Escreve em `teste_imoveis` — mesmos nomes de coluna que `imoveis`, todas
`text` (staging, sem risco de falha de cast) + `extra` jsonb para
cabeçalhos do relatório sem correspondência. NÃO tocar em `imoveis`
(produção) enquanto for teste.

Correr a partir de backend/: python scripts/export_relatorio_imoveis.py
Depende de backend/scripts/requirements-scraper.txt (playwright, openpyxl)
+ `python -m playwright install chromium` (uma vez).
"""
import asyncio
import csv
import datetime
import io
import re
import sys
import unicodedata
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl
from playwright.async_api import async_playwright

from app.config import settings
from app.db.supabase_client import get_supabase
from app.integrations import egorealestate_crm

REPORT_NAME = "jmarques_imoveis"
STATUS_FILTER = "Disponível"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"

# Mesmos nomes que `imoveis` (migration 0009) — sem `id`/`extra`/`criado_em`.
_KNOWN_COLUMNS = {
    "imovel_ref", "natureza", "disponibilidade", "estado", "fonte", "titulo", "descricao",
    "proprietario", "angariador", "vendedor", "quartos", "casas_banho", "suites", "piso",
    "num_pisos", "numero", "fracao", "area_util", "area_bruta", "area_terreno", "conservacao",
    "certificacao_energetica", "venda_preco", "arrendamento_preco", "comissao_agencia",
    "comissao_angariador", "comissao_vendedor", "exclusividade", "morada", "codigo_postal",
    "concelho", "freguesia", "zona", "piscina", "garagem", "jardim", "terraco", "varanda",
    "vista_mar", "vista_praia", "ar_condicionado", "elevador", "aquecimento_central",
    "arrecadacao", "estacionamento", "portais", "foto_principal", "fotos", "panoramic_url",
    "video_url", "ego_id", "ego_atualizado_em", "data_criacao", "data_alteracao",
}

# O relatório devolve 1 linha por (imóvel × nota) — dados do imóvel repetem-se
# idênticos em cada linha, só estes 4 campos mudam por nota. `_merge_notas`
# funde-os num só registo por `imovel_ref`.
_NOTA_FIELDS = ("Criado em", "Criado por", "Tipo de nota", "Anexos")

# Cabeçalhos reais do relatório "jmarques_imoveis_notas" (visto 2026-07-26) que
# não batem por normalização simples — mapeados à mão depois de inspeccionar
# um ficheiro real. Comissões com split Venda/Arrendamento ficam de fora de
# propósito: `imoveis` só tem 1 coluna p/ cada, escolher um dos dois perderia
# informação silenciosamente — melhor deixar em `extra`.
_ALIASES = {
    "referencia": "imovel_ref",
    "casa_s_de_banho": "casas_banho",
    "suite_s": "suites",
    "numero_de_pisos": "num_pisos",
    "venda": "venda_preco",
    "arrendamento": "arrendamento_preco",
    "varandas": "varanda",
    "vista_para_mar": "vista_mar",
    "vista_para_praia": "vista_praia",
    "publicacao_para_site_portais": "portais",
    "data_de_criacao": "data_criacao",
    "data_de_alteracao": "data_alteracao",
    "contrato_de_mediacao_exclusividade": "exclusividade",
}


def _normalize(header: str) -> str:
    text = unicodedata.normalize("NFKD", header).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_").lower()


def _dedupe_headers(headers: list[str]) -> list[str]:
    """O relatório repete nomes de coluna (ex: "Área útil" 2x) — sem isto
    `dict(zip(header, row))` perdia a 1ª ocorrência silenciosamente."""
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
        result.append(h if seen[h] == 1 else f"{h} ({seen[h]})")
    return result


def _map_row(row: dict) -> dict:
    record: dict = {"extra": {}}
    for header, value in row.items():
        key = _normalize(str(header))
        key = _ALIASES.get(key, key)
        if key in _KNOWN_COLUMNS and key not in record:
            record[key] = None if value is None else str(value)
        else:
            record["extra"][str(header)] = value
    return record


def _merge_notas(records: list[dict]) -> list[dict]:
    grouped: dict[str, dict] = {}
    order: list[str] = []
    ignorados = 0
    for record in records:
        ref = record.get("imovel_ref")
        if not ref:
            # sem referência (ex: registos de "Chave" avulsos no relatório,
            # sem ligação a um imóvel) — ignorar, não são imóveis.
            ignorados += 1
            continue
        is_new = ref not in grouped
        if is_new:
            grouped[ref] = {k: v for k, v in record.items() if k != "extra"}
            grouped[ref]["extra"] = {k: v for k, v in record["extra"].items() if k not in _NOTA_FIELDS}
            grouped[ref]["extra"]["notas"] = []
            order.append(ref)
        else:
            # ref repetida por outro motivo que não notas — bug conhecido
            # do eGO (mesma Reference em propriedades distintas, ex: FH2460 4D
            # com area_bruta/piso diferentes). Não há forma de adivinhar qual
            # fica — mantém a 1ª, avisa em vez de perder em silêncio.
            print(f'  aviso: "{ref}" repetido no relatório com dados possivelmente diferentes (bug eGO conhecido) — mantida só a 1ª ocorrência')
        nota = {k: record["extra"].get(k) for k in _NOTA_FIELDS}
        if any(v is not None for v in nota.values()):
            grouped[ref]["extra"]["notas"].append(nota)
    if ignorados:
        print(f"  {ignorados} linha(s) sem imovel_ref ignoradas (não são imóveis)")
    return [grouped[ref] for ref in order]


async def _httpx_cookies() -> list[dict]:
    async with egorealestate_crm.authenticated_client() as client:
        await egorealestate_crm._login(client)
        return [
            {"name": c.name, "value": c.value, "domain": c.domain, "path": c.path or "/"}
            for c in client.cookies.jar
        ]


def _json_safe(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def _load_workbook_safe(path: Path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except IndexError:
        # eGO exporta xlsx com `styles.xml` a referenciar cellStyleXfs fora de
        # gama (quirk comum de geradores .NET tipo ClosedXML/EPPlus) — openpyxl
        # parte ao juntar named styles. Não precisamos de estilos, só dados:
        # stripar <cellStyles> e tentar de novo numa cópia patched.
        patched = path.with_suffix(".patched.xlsx")
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    # elementos vêm com prefixo de namespace (ex: <x:cellStyles>)
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*>.*?</(\w+:)?cellStyles>", b"", data, flags=re.DOTALL)
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*/>", b"", data)
                zout.writestr(item, data)
        return openpyxl.load_workbook(patched, data_only=True)


def _parse_rows(path: Path) -> list[dict]:
    if path.suffix.lower() == ".xlsx":
        wb = _load_workbook_safe(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = _dedupe_headers([str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])])
        return [{k: _json_safe(v) for k, v in zip(header, row)} for row in rows[1:]]
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        return list(csv.DictReader(io.StringIO(text)))
    raise ValueError(f"Formato não suportado: {path.suffix}")


async def main() -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    print("A obter sessão autenticada (httpx)...")
    cookies = await _httpx_cookies()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=False, slow_mo=400)
        context = await browser.new_context(accept_downloads=True)
        await context.add_cookies(cookies)
        page = await context.new_page()

        download_path: Path | None = None
        download_done = asyncio.Event()

        async def _on_download(download):
            nonlocal download_path
            download_path = OUTPUT_DIR / download.suggested_filename
            await download.save_as(download_path)
            print(f"Download: {download_path}")
            download_done.set()

        context.on("download", _on_download)

        print("A navegar para Imóveis...")
        await page.goto(
            f"{settings.egorealestate_crm_base_url}/egocore/realestates",
            wait_until="networkidle",
            timeout=30000,
        )

        if "login" in page.url:
            raise RuntimeError("Sessão não autenticou — cookies inválidas ou login mudou.")

        await page.wait_for_timeout(2000)

        print(f'A aplicar filtro "{STATUS_FILTER}"...')
        clicked = await page.evaluate(
            """(filtro) => {
                const span = Array.from(document.querySelectorAll('span.sideTag'))
                    .find(e => e.textContent.trim() === filtro);
                if (!span) return false;
                const link = span.querySelector('a') || span;
                link.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
                return true;
            }""",
            STATUS_FILTER,
        )
        if not clicked:
            raise RuntimeError(f'Filtro "{STATUS_FILTER}" não encontrado na sidebar.')
        await page.wait_for_timeout(2000)

        print("A limpar selecção...")
        await page.evaluate("document.querySelector('.btDeselectAll, a.btDeselectAll')?.click()")
        await page.wait_for_timeout(1000)

        print("A seleccionar todos...")
        await page.evaluate("document.querySelector('.btSelectAll, a.btSelectAll')?.click()")
        await page.wait_for_timeout(1500)

        await page.evaluate(
            """() => {
                const ok = Array.from(document.querySelectorAll('button, a'))
                    .find(b => b.textContent.trim() === 'OK' || b.textContent.trim() === 'Fechar');
                if (ok) ok.click();
            }"""
        )
        await page.wait_for_timeout(500)

        print("A abrir Relatórios...")
        clicked_rel = await page.evaluate(
            """() => {
                const el = Array.from(document.querySelectorAll('a'))
                    .find(e => e.textContent.trim() === 'Relatórios' && e.getAttribute('onclick')?.includes('popupreports'));
                if (el) { el.click(); return true; }
                return false;
            }"""
        )
        if not clicked_rel:
            raise RuntimeError('Link "Relatórios" não encontrado — selecção pode ter falhado.')
        await page.wait_for_timeout(3000)

        print(f'A clicar no relatório "{REPORT_NAME}"...')
        clicked_report = await page.evaluate(
            """(nome) => {
                const items = Array.from(document.querySelectorAll('#ReportsList .popupReportItem, #ReportsPopup .popupReportItem'));
                const item = items.find(el => el.textContent.trim() === nome);
                if (item) { (item.querySelector('a') || item).click(); return true; }
                return false;
            }""",
            REPORT_NAME,
        )
        if not clicked_report:
            raise RuntimeError(f'Relatório "{REPORT_NAME}" não encontrado na lista de relatórios gravados.')

        try:
            await asyncio.wait_for(download_done.wait(), timeout=30)
        except asyncio.TimeoutError:
            pass

        await browser.close()

    if not download_done.is_set() or download_path is None:
        raise RuntimeError("Nenhum download detectado em 30s.")

    print(f"A parsear {download_path.name}...")
    rows = _parse_rows(download_path)
    print(f"{len(rows)} linhas no relatório.")
    if not rows:
        return

    registos = _merge_notas([_map_row(row) for row in rows])
    print(f"{len(registos)} imóveis únicos (notas fundidas).")

    print("A gravar em teste_imoveis...")
    supabase = get_supabase()
    for i in range(0, len(registos), 100):
        supabase.table("teste_imoveis").insert(registos[i : i + 100]).execute()
    print(f"{len(registos)} linhas gravadas em teste_imoveis.")
    sem_match = {k for r in registos for k in r["extra"]}
    if sem_match:
        print(f"Colunas do relatório sem correspondência em `imoveis` (foram para `extra`): {sorted(sem_match)}")


if __name__ == "__main__":
    asyncio.run(main())
