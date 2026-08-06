"""Scraper Playwright — relatório "jmarques_todas_as_colunas" do eGO CRM
(módulo Oportunidades), filtro "Editado em > Últimas 48 horas". Mesmo
mecanismo que `backend/scripts/export_relatorio_oportunidades.py` (login
httpx + cookies injectadas, popup Relatórios) — ver esse ficheiro para o
detalhe do porquê de cada passo de navegação/filtro.

Download: dispara o relatório e lê a URL directa do ficheiro (.xlsx em
media.egorealestate.com) do corpo da resposta JSON de
`POST /egocore/report/export` — NÃO depende do popup que a página abre nem
do evento `download` do browser. Confirmado ao vivo: em Fly.io esse popup
fica sempre em branco e nunca navega para o ficheiro (falha 100%
reproduzível, nunca vista em dev local — causa exacta do lado do JS da
eGO desconhecida, mas irrelevante, a resposta já traz tudo p/ ignorar o
popup por completo).

Depois de mapear/agrupar (`mapping_todas_colunas`), grava directo em
produção (`upsert`) — `oportunidades`/`notas`/`tarefas`/`contactos`.

Correr localmente a partir de scraper/: python oportunidades_completo.py
Depende de scraper/requirements.txt + `python -m playwright install chromium`
(uma vez).
"""
import asyncio
import csv
import datetime
import io
import re
import zipfile
from pathlib import Path

import httpx
import openpyxl
from playwright.async_api import async_playwright

import config
import ego_auth

REPORT_NAME = "jmarques_todas_as_colunas"
OUTPUT_DIR = Path(__file__).resolve().parent / "output"


def _json_safe(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def _load_workbook_safe(path: Path):
    try:
        return openpyxl.load_workbook(path, data_only=True)
    except IndexError:
        # eGO exporta xlsx com `styles.xml` a referenciar cellStyleXfs fora de
        # gama (quirk comum de geradores .NET) — ver export_relatorio_imoveis.py.
        patched = path.with_suffix(".patched.xlsx")
        with zipfile.ZipFile(path, "r") as zin, zipfile.ZipFile(patched, "w") as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*>.*?</(\w+:)?cellStyles>", b"", data, flags=re.DOTALL)
                    data = re.sub(rb"<(\w+:)?cellStyles[^>]*/>", b"", data)
                zout.writestr(item, data)
        return openpyxl.load_workbook(patched, data_only=True)


def _dedupe_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    result = []
    for h in headers:
        seen[h] = seen.get(h, 0) + 1
        result.append(h if seen[h] == 1 else f"{h} ({seen[h]})")
    return result


def _parse_rows(path: Path) -> tuple[list[str], list[dict]]:
    if path.suffix.lower() == ".xlsx":
        wb = _load_workbook_safe(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], []
        header = _dedupe_headers([str(h).strip() if h else f"col{i}" for i, h in enumerate(rows[0])])
        return header, [{k: _json_safe(v) for k, v in zip(header, row)} for row in rows[1:]]
    if path.suffix.lower() == ".csv":
        text = path.read_text(encoding="utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
        return (list(rows[0].keys()) if rows else []), rows
    raise ValueError(f"Formato não suportado: {path.suffix}")


async def _trigger_and_download(headless: bool = True) -> Path:
    """Dispara o relatório e descarrega o .xlsx via httpx a partir da URL
    devolvida pelo próprio POST /egocore/report/export.

    Não depende do popup que a página abre nem do evento `download` do
    browser — confirmado ao vivo (Fly.io) que esse popup fica sempre em
    branco e nunca navega (falha silenciosa do lado do JS da página,
    reprodutível 100% das vezes em Fly.io, nunca reproduzida em dev local —
    causa exacta desconhecida, mas irrelevante: a resposta JSON de
    `/report/export` já traz a URL directa do ficheiro em `data`, então
    saltamos o popup por completo e descarregamos nós mesmos."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    print("A obter sessão autenticada (httpx)...")
    cookies = await ego_auth.session_cookies()

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(headless=headless)
        context = await browser.new_context(accept_downloads=True)
        await context.add_cookies(cookies)
        page = await context.new_page()

        export_url_future: asyncio.Future[str] = asyncio.get_event_loop().create_future()

        async def _on_response(resp):
            if "/report/export" in resp.url and not export_url_future.done():
                try:
                    body = await resp.json()
                except Exception:
                    return
                url = body.get("data")
                if url and isinstance(url, str) and url.startswith("http"):
                    export_url_future.set_result(url)

        context.on("response", _on_response)

        print("A navegar para Oportunidades...")
        await page.goto(
            f"{config.egorealestate_crm_base_url}/egocore/leads",
            wait_until="networkidle",
            timeout=30000,
        )
        if "login" in page.url:
            raise RuntimeError("Sessão não autenticou — cookies inválidas ou login mudou.")
        await page.wait_for_timeout(2000)

        # "Todas as opções disponíveis" refere-se às COLUNAS do relatório
        # (jmarques_todas_as_colunas tem todas, ao contrário de
        # jmarques_oportunidades_notas), não a um pull sem filtro de tempo —
        # confirmado ao vivo: sem período, o eGO não devolve download directo
        # (portefólio grande demais), manda antes por email. Por isso aplica-se
        # o MESMO filtro "Últimas 48 horas" do scraper existente
        # (export_relatorio_oportunidades.py) — mesmo padrão de desselecção de
        # "Minhas oportunidades" (clicar na tag já seleccionada, "Todas as
        # oportunidades" não resolve).
        print('A desseleccionar "Minhas oportunidades"...')
        removeu = await page.evaluate(
            """() => {
                const els = Array.from(document.querySelectorAll('span.sideTag a'))
                    .filter(e => e.textContent.trim() === 'Minhas oportunidades');
                if (!els[1]) return false;
                els[1].dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
                return true;
            }"""
        )
        if not removeu:
            raise RuntimeError('"Minhas oportunidades" não encontrado para desseleccionar.')
        await page.wait_for_timeout(2000)

        print('A aplicar filtro "Editado em > Últimas 48 horas"...')
        aplicou_periodo = await page.evaluate(
            """() => {
                const span = Array.from(document.querySelectorAll('span.sideTag'))
                    .find(e => e.textContent.trim() === 'Últimas 48 horas');
                if (!span) return false;
                const link = span.querySelector('a') || span;
                link.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
                return true;
            }"""
        )
        if not aplicou_periodo:
            raise RuntimeError('Filtro "Últimas 48 horas" não encontrado.')
        await page.wait_for_timeout(2000)

        print("A limpar selecção...")
        await page.evaluate("document.querySelector('.btDeselectAll, a.btDeselectAll')?.click()")
        await page.wait_for_timeout(1000)

        print("A seleccionar todos...")
        await page.evaluate("document.querySelector('.btSelectAll, a.btSelectAll')?.click()")
        await page.wait_for_timeout(1500)

        await page.evaluate(
            """() => {
                const ok = Array.from(document.querySelectorAll('button, a'))
                    .find(b => b.textContent.trim() === 'OK' || b.textContent.trim() === 'Fechar');
                if (ok) ok.click();
            }"""
        )
        await page.wait_for_timeout(500)

        print("A abrir Relatórios...")
        clicked_rel = await page.evaluate(
            """() => {
                const el = Array.from(document.querySelectorAll('a'))
                    .find(e => e.textContent.trim() === 'Relatórios' && e.getAttribute('onclick')?.includes('popupreports'));
                if (el) { el.click(); return true; }
                return false;
            }"""
        )
        if not clicked_rel:
            raise RuntimeError('Link "Relatórios" não encontrado — selecção pode ter falhado (0 resultados?).')
        await page.wait_for_timeout(3000)

        print(f'A clicar no relatório "{REPORT_NAME}"...')
        clicked_report = await page.evaluate(
            """(nome) => {
                const items = Array.from(document.querySelectorAll('#ReportsList .popupReportItem, #ReportsPopup .popupReportItem'));
                const item = items.find(el => el.textContent.trim() === nome);
                if (item) { (item.querySelector('a') || item).click(); return true; }
                return false;
            }""",
            REPORT_NAME,
        )
        if not clicked_report:
            raise RuntimeError(f'Relatório "{REPORT_NAME}" não encontrado na lista de relatórios gravados.')

        try:
            file_url = await asyncio.wait_for(export_url_future, timeout=90)
        except asyncio.TimeoutError:
            raise RuntimeError(
                "Relatório não devolveu URL de download em 90s — se o filtro (48h) devolver muitos "
                "resultados, o eGO pode enviar o relatório por email em vez de gerar download directo."
            )

        await browser.close()

    # URL já vem assinada (query param "V=...") em media.egorealestate.com —
    # domínio diferente do admin, não precisa das cookies de sessão.
    print(f"A descarregar {file_url}...")
    async with httpx.AsyncClient(timeout=60, follow_redirects=True) as client:
        resp = await client.get(file_url)
        resp.raise_for_status()
        download_path = OUTPUT_DIR / file_url.split("/")[-1].split("?")[0]
        download_path.write_bytes(resp.content)
    print(f"Download: {download_path}")
    return download_path


async def download_and_parse(headless: bool = True) -> tuple[Path, list[str], list[dict]]:
    download_path = await _trigger_and_download(headless=headless)
    print(f"A parsear {download_path.name}...")
    header, rows = _parse_rows(download_path)
    print(f"{len(header)} colunas, {len(rows)} linhas no relatório.")
    return download_path, header, rows


async def run(headless: bool = True) -> dict:
    """Fluxo completo: dispara relatório, faz download, mapeia e agrupa
    (`mapping_todas_colunas`), grava em produção (`upsert.run`). Devolve o
    resumo por tabela."""
    import mapping_todas_colunas as m
    import upsert
    from supabase import create_client

    import config

    _, _, rows = await download_and_parse(headless=headless)
    classified = [m.classify(m.map_row(r)) for r in rows]
    batch = m.group(classified)
    print(
        f"oportunidades={len(batch['oportunidades'])} notas={len(batch['notas'])} "
        f"tarefas={len(batch['tarefas'])} prefs={len(batch['prefs'])} contactos={len(batch['contactos'])}"
    )

    supabase = create_client(config.supabase_imoveis_url, config.supabase_imoveis_key)
    return upsert.run(supabase, batch)


if __name__ == "__main__":
    resumo = asyncio.run(run(headless=False))
    print("\nRESUMO:", resumo)
